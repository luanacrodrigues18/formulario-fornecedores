from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CABECALHOS = [
    "ID",
    "Hora de início",
    "Hora da conclusão",
    "Email",
    "Nome",
    "Número do PO com Release",
    "Data da Promessa",
    "Observações de Coleta",
    "Número da NF",
    "Informe o número da linha",
]

EXEMPLOS = [
    [
        1,
        "17/06/2026 09:15:00",
        "17/06/2026 09:18:30",
        "fornecedor1@empresa.com",
        "João Silva",
        "PO-2026-001-R1",
        "20/06/2026",
        "Coleta agendada para manhã",
        "5862",
        "10",
    ],
    [
        2,
        "17/06/2026 10:00:00",
        "17/06/2026 10:05:12",
        "fornecedor2@empresa.com",
        "Maria Santos",
        "PO-2026-002-R2",
        "25/06/2026",
        "Aguardando confirmação",
        "1234",
        "15",
    ],
]

LARGURAS = [8, 22, 22, 30, 25, 28, 18, 35, 18, 22]

ARQUIVO_SAIDA = "template_fornecedores.xlsx"


def gerar_template(caminho: str = ARQUIVO_SAIDA) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")

    for col_idx, titulo in enumerate(CABECALHOS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, linha in enumerate(EXEMPLOS, start=2):
        for col_idx, valor in enumerate(linha, start=1):
            ws.cell(row=row_idx, column=col_idx, value=valor)

    for idx, largura in enumerate(LARGURAS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = largura

    ws.freeze_panes = "A2"
    wb.save(caminho)
    return caminho


if __name__ == "__main__":
    arquivo = gerar_template()
    print(f"Template gerado com sucesso: {arquivo}")
