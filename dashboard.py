from datetime import date, datetime, timedelta
from io import BytesIO
import os

import altair as alt
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from database import (
    COLUNAS_EXIBICAO,
    agora_brasil,
    buscar_todos,
    criar_tabela,
    formatar_datetime,
    parse_datetime,
    registro_incompleto,
    status_registro,
    supabase_configurado,
    valor_vazio,
)
from planilha import buscar_por_po_e_linha, montar_link_formulario, exportar_retorno_fup_excel
from prazo_resposta import (
    carregar_prazo,
    contar_por_prazo,
    data_limite_atual,
    data_limite_em_dias,
    garantir_arquivo_prazo,
    limpar_prazo,
    salvar_prazo,
    status_prazo_envio,
)
from pathlib import Path

st.set_page_config(
    page_title="Dashboard - Fornecedores",
    page_icon="📊",
    layout="wide",
)

COLUNAS_ORDEM = list(COLUNAS_EXIBICAO.keys())
COL_STATUS = "Status"
COL_PRAZO = "Prazo do form"
LARGURAS_COLUNAS = [8, 18, 22, 22, 30, 25, 28, 18, 35, 18, 22]
COL_NF = COLUNAS_EXIBICAO["numero_nf"]
COL_OBS = COLUNAS_EXIBICAO["observacoes_coleta"]
OPCOES_PAGINA = [10, 25, 50, 100]
FORM_BASE_URL_PADRAO = os.getenv("FORM_BASE_URL", "http://localhost:8501")


def _parse_datetime(valor) -> datetime | None:
    return parse_datetime(valor)


def _datetime_fallback() -> datetime:
    return datetime.min.replace(tzinfo=agora_brasil().tzinfo)


def _valor_vazio(valor) -> bool:
    return valor_vazio(valor)


def ordenar_por_data(registros: list[dict]) -> list[dict]:
    return sorted(
        registros,
        key=lambda r: _parse_datetime(r.get("hora_conclusao")) or _datetime_fallback(),
        reverse=True,
    )


def ultimo_envio(registros: list[dict]) -> datetime | None:
    datas = [_parse_datetime(r.get("hora_conclusao")) for r in registros]
    datas_validas = [d for d in datas if d]
    return max(datas_validas) if datas_validas else None


def contar_envios_hoje(registros: list[dict]) -> int:
    hoje = agora_brasil().date()
    return sum(
        1
        for r in registros
        if (_parse_datetime(r.get("hora_conclusao")) or _datetime_fallback()).date() == hoje
    )


def contar_envios_semana(registros: list[dict]) -> int:
    inicio = agora_brasil().date() - timedelta(days=6)
    return sum(
        1
        for r in registros
        if (_parse_datetime(r.get("hora_conclusao")) or _datetime_fallback()).date() >= inicio
    )


def contar_sem_nf(registros: list[dict]) -> int:
    return sum(1 for r in registros if _valor_vazio(r.get("numero_nf")))


def registros_para_dataframe(registros: list[dict]) -> pd.DataFrame:
    colunas = [COLUNAS_EXIBICAO[c] for c in COLUNAS_ORDEM]
    colunas_tabela = [colunas[0], COL_STATUS, COL_PRAZO] + colunas[1:]
    limite = data_limite_atual()
    if not registros:
        return pd.DataFrame(columns=colunas_tabela)

    linhas = []
    for registro in registros:
        linha = {
            COL_STATUS: status_registro(registro),
            COL_PRAZO: status_prazo_envio(registro, limite),
        }
        for campo in COLUNAS_ORDEM:
            valor = registro.get(campo, "")
            if campo in ("hora_inicio", "hora_conclusao"):
                valor = formatar_datetime(valor)
            elif campo == "data_promessa" and valor:
                dt = _parse_datetime(valor)
                valor = dt.strftime("%d/%m/%Y") if dt else valor
            linha[COLUNAS_EXIBICAO[campo]] = valor
        linhas.append(linha)

    return pd.DataFrame(linhas, columns=colunas_tabela)


def aplicar_destaque(df: pd.DataFrame):
    def highlight(row):
        nf = row.get(COL_NF, "")
        obs = row.get(COL_OBS, "")
        if _valor_vazio(nf) or _valor_vazio(obs):
            return ["background-color: #fef9c3"] * len(row)
        return [""] * len(row)

    return df.style.apply(highlight, axis=1)


def gerar_excel(registros: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Fornecedores"

    cabecalhos = [COLUNAS_EXIBICAO[c] for c in COLUNAS_ORDEM]
    cabecalhos_excel = [cabecalhos[0], COL_STATUS, COL_PRAZO] + cabecalhos[1:]
    fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    limite = data_limite_atual()

    for col_idx, titulo in enumerate(cabecalhos_excel, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, registro in enumerate(registros, start=2):
        ws.cell(row=row_idx, column=1, value=registro.get("id", ""))
        ws.cell(row=row_idx, column=2, value=status_registro(registro))
        ws.cell(row=row_idx, column=3, value=status_prazo_envio(registro, limite))
        for col_idx, campo in enumerate(COLUNAS_ORDEM[1:], start=4):
            valor = registro.get(campo, "")
            if campo in ("hora_inicio", "hora_conclusao"):
                valor = formatar_datetime(valor)
            ws.cell(row=row_idx, column=col_idx, value=valor)

    for idx, largura in enumerate([8, 14, 14] + LARGURAS_COLUNAS[1:], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = largura

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def aplicar_filtros(
    registros: list[dict],
    filtro_nome: str,
    filtro_email: str,
    filtro_po: str,
    filtro_nf: str,
    filtro_linha: str,
    busca_geral: str,
    filtro_data,
    somente_incompletos: bool = False,
    filtro_prazo: str = "Todos",
) -> list[dict]:
    resultado = registros
    limite = data_limite_atual()

    if somente_incompletos:
        resultado = [r for r in resultado if registro_incompleto(r)]

    if filtro_prazo and filtro_prazo != "Todos":
        resultado = [
            r for r in resultado if status_prazo_envio(r, limite) == filtro_prazo
        ]

    if filtro_nome.strip():
        termo = filtro_nome.strip().lower()
        resultado = [r for r in resultado if termo in str(r.get("nome", "")).lower()]

    if filtro_email.strip():
        termo = filtro_email.strip().lower()
        resultado = [r for r in resultado if termo in str(r.get("email", "")).lower()]

    if filtro_po.strip():
        termo = filtro_po.strip().lower()
        resultado = [
            r for r in resultado if termo in str(r.get("numero_po_com_release", "")).lower()
        ]

    if filtro_nf.strip():
        termo = filtro_nf.strip().lower()
        resultado = [r for r in resultado if termo in str(r.get("numero_nf", "")).lower()]

    if filtro_linha.strip():
        termo = filtro_linha.strip().lower()
        resultado = [r for r in resultado if termo in str(r.get("numero_linha", "")).lower()]

    if busca_geral.strip():
        termo = busca_geral.strip().lower()
        resultado = [
            r
            for r in resultado
            if any(termo in str(r.get(campo, "")).lower() for campo in COLUNAS_ORDEM)
        ]

    if filtro_data is not None:
        data_filtro = filtro_data.isoformat()
        resultado = [
            r for r in resultado if str(r.get("data_promessa", ""))[:10] == data_filtro
        ]

    return resultado


def grafico_envios_por_dia(registros: list[dict]) -> pd.DataFrame:
    hoje = agora_brasil().date()
    dias = [hoje - timedelta(days=i) for i in range(6, -1, -1)]
    contagem = {d.strftime("%d/%m"): 0 for d in dias}

    for registro in registros:
        dt = _parse_datetime(registro.get("hora_conclusao"))
        if dt and dt.date() in dias:
            contagem[dt.strftime("%d/%m")] += 1

    return pd.DataFrame({"Dia": list(contagem.keys()), "Envios": list(contagem.values())})


def grafico_top_fornecedores(registros: list[dict]) -> pd.DataFrame:
    if not registros:
        return pd.DataFrame(columns=["Fornecedor", "Fornecedor_curto", "Envios"])

    series = pd.Series([str(r.get("nome", "—")) for r in registros])
    top = series.value_counts().head(10).reset_index()
    top.columns = ["Fornecedor", "Envios"]
    top["Fornecedor_curto"] = top["Fornecedor"].apply(
        lambda nome: (nome[:42] + "…") if len(nome) > 42 else nome
    )
    return top.sort_values("Envios", ascending=True)


def _chart_envios_por_dia(df: pd.DataFrame) -> alt.Chart:
    max_envios = max(df["Envios"].max(), 1)
    base = alt.Chart(df).encode(
        x=alt.X("Dia:N", title="Dia", sort=None, axis=alt.Axis(labelAngle=0)),
        y=alt.Y(
            "Envios:Q",
            title="Envios",
            scale=alt.Scale(domain=[0, max_envios * 1.2]),
        ),
        tooltip=[alt.Tooltip("Dia", title="Dia"), alt.Tooltip("Envios", title="Envios")],
    )
    barras = base.mark_bar(color="#2d5a8e", cornerRadiusTopLeft=4, cornerRadiusTopRight=4)
    rotulos = base.mark_text(dy=-6, color="#1e3a5f", fontSize=12, fontWeight="bold").encode(
        text=alt.Text("Envios:Q")
    )
    return (barras + rotulos).properties(height=300)


def _chart_top_fornecedores(df: pd.DataFrame) -> alt.Chart:
    altura = max(300, len(df) * 32)
    max_envios = max(df["Envios"].max(), 1) if not df.empty else 1
    base = alt.Chart(df).encode(
        y=alt.Y(
            "Fornecedor_curto:N",
            title="Fornecedor",
            sort=alt.EncodingSortField(field="Envios", order="descending"),
        ),
        x=alt.X(
            "Envios:Q",
            title="Envios",
            scale=alt.Scale(domain=[0, max_envios * 1.25]),
        ),
        tooltip=[
            alt.Tooltip("Fornecedor", title="Fornecedor"),
            alt.Tooltip("Envios", title="Envios"),
        ],
    )
    barras = base.mark_bar(color="#2d5a8e", cornerRadiusTopRight=4, cornerRadiusBottomRight=4)
    rotulos = base.mark_text(
        align="left",
        baseline="middle",
        dx=5,
        color="#1e3a5f",
        fontSize=12,
        fontWeight="bold",
    ).encode(text=alt.Text("Envios:Q"))
    return (barras + rotulos).properties(height=altura)


# ── Cabeçalho ───────────────────────────────────────────────────────────────
col_titulo, col_btn1, col_btn2 = st.columns([4, 1, 1])
with col_titulo:
    st.title("📊 Dashboard de Fornecedores")
with col_btn1:
    st.write("")
    if st.button("🔄 Atualizar", type="primary", use_container_width=True):
        st.rerun()
with col_btn2:
    st.write("")
    auto_atualizar = st.checkbox("Auto 60s", help="Atualiza automaticamente a cada 60 segundos")

if auto_atualizar:
    components.html(
        "<script>setTimeout(function(){window.parent.location.reload();}, 60000);</script>",
        height=0,
    )

agora = agora_brasil()
st.caption(
    f"Última leitura: {agora.strftime('%d/%m/%Y %H:%M:%S')} (horário de Brasília) · "
    "Linhas em amarelo = sem NF ou sem observação"
)

try:
    garantir_arquivo_prazo()
except Exception:
    pass

if supabase_configurado():
    try:
        criar_tabela()
        st.caption("Dados carregados do Supabase.")
    except RuntimeError as exc:
        st.warning(f"Supabase com erro: {exc}. Tentando exibir dados do Excel.")
else:
    st.info("Exibindo respostas salvas em `formulario_respostas.xlsx`.")

try:
    registros = ordenar_por_data(buscar_todos())
except Exception as exc:
    st.error(f"Erro ao buscar registros: {exc}")
    st.stop()

ultimo = ultimo_envio(registros)
limite_atual = data_limite_atual()
contagem_prazo = contar_por_prazo(registros, limite_atual)

# ── Cards ───────────────────────────────────────────────────────────────────
c1, c2, c3, c4, c5 = st.columns(5)
c1.metric("Total geral", len(registros))
c2.metric("Envios hoje", contar_envios_hoje(registros))
c3.metric("Últimos 7 dias", contar_envios_semana(registros))
c4.metric("Sem Número da NF", contar_sem_nf(registros))
c5.metric(
    "Último envio",
    ultimo.strftime("%d/%m %H:%M") if ultimo else "—",
)

if limite_atual:
    p1, p2, p3 = st.columns(3)
    p1.metric("Retornos no prazo", contagem_prazo.get("No prazo", 0))
    p2.metric("Retornos atrasados", contagem_prazo.get("Atrasado", 0))
    p3.metric("Data limite do form", limite_atual.strftime("%d/%m/%Y"))

# ── Prazo para preencher o formulário ───────────────────────────────────────
with st.expander("⏰ Prazo para o fornecedor preencher o formulário", expanded=not limite_atual):
    st.markdown(
        """
        Defina até quando o fornecedor deve **abrir o formulário e enviar o retorno**.
        No retorno, o fornecedor informa a **Data da Promessa** (quando o material estará pronto).

        - **X dias** → calcula a data limite a partir de hoje e grava essa data fixa  
        - **Data limite** → você escolhe o dia final  
        """
    )
    cfg = carregar_prazo()
    if limite_atual:
        obs_atual = (cfg or {}).get("observacao") or ""
        st.info(
            f"Prazo atual: **{limite_atual.strftime('%d/%m/%Y')}**"
            + (f" — {obs_atual}" if obs_atual else "")
        )

    modo = st.radio(
        "Como definir",
        ["Em X dias (a partir de hoje)", "Data limite fixa"],
        horizontal=True,
        key="modo_prazo",
    )
    obs = st.text_input(
        "Observação (opcional, aparece no formulário)",
        value=(cfg or {}).get("observacao", "") if cfg else "",
        placeholder="Ex.: campanha FUP julho",
        key="obs_prazo",
    )

    col_a, col_b, col_c = st.columns([2, 1, 1])
    with col_a:
        if modo.startswith("Em X"):
            dias_prazo = st.number_input(
                "Dias para responder",
                min_value=0,
                max_value=365,
                value=int((cfg or {}).get("dias_origem") or 7),
                step=1,
                key="dias_prazo",
            )
            preview = data_limite_em_dias(int(dias_prazo))
            st.caption(f"Data limite calculada: **{preview.strftime('%d/%m/%Y')}**")
        else:
            valor_data = limite_atual or (agora.date() + timedelta(days=7))
            data_escolhida = st.date_input(
                "Data limite",
                value=valor_data,
                format="DD/MM/YYYY",
                key="data_prazo",
            )
            dias_prazo = None
            preview = data_escolhida

    with col_b:
        st.write("")
        st.write("")
        if st.button("Salvar prazo", type="primary", use_container_width=True):
            try:
                salvar_prazo(
                    preview,
                    dias_origem=int(dias_prazo) if dias_prazo is not None else None,
                    observacao=obs,
                )
                st.success(f"Prazo salvo: {preview.strftime('%d/%m/%Y')}")
                st.rerun()
            except Exception as exc:
                st.error(f"Erro ao salvar prazo: {exc}")
    with col_c:
        st.write("")
        st.write("")
        if st.button("Remover prazo", use_container_width=True, disabled=not limite_atual):
            limpar_prazo()
            st.rerun()

# ── Gráficos ────────────────────────────────────────────────────────────────
if registros:
    g1, g2 = st.columns(2)
    with g1:
        st.subheader("Envios por dia (7 dias)")
        df_dias = grafico_envios_por_dia(registros)
        st.altair_chart(_chart_envios_por_dia(df_dias), use_container_width=True)
    with g2:
        st.subheader("Top 10 fornecedores")
        df_top = grafico_top_fornecedores(registros)
        st.altair_chart(_chart_top_fornecedores(df_top), use_container_width=True)

# ── Gerador de link ─────────────────────────────────────────────────────────
with st.expander("🔗 Gerar link para fornecedor", expanded=False):
    lg1, lg2, lg3 = st.columns([2, 1, 2])
    with lg1:
        link_po = st.text_input("PO com Release", key="gerador_link_po", placeholder="4133600-23")
    with lg2:
        link_linha = st.text_input("Linha", key="gerador_link_linha", placeholder="17")
    with lg3:
        form_base_url = st.text_input(
            "URL do formulário",
            value=FORM_BASE_URL_PADRAO,
            help="Em produção, use a URL pública do formulário.",
        )

    if link_po.strip() and link_linha.strip():
        url_fornecedor = montar_link_formulario(form_base_url, link_po, link_linha)
        st.code(url_fornecedor, language=None)
        pedidos = buscar_por_po_e_linha(link_po.strip(), link_linha.strip())
        if pedidos:
            fornecedor = pedidos[0].get("fornecedor", "—")
            st.caption(f"Pedido encontrado no FUP: **{fornecedor}** — copie o link e envie por e-mail.")
        else:
            st.warning("PO/linha não encontrados no FUP. Verifique os dados antes de enviar o link.")

# ── Exportar retorno (sem alterar o .xlsm) ───────────────────────────────────
with st.expander("📥 Exportar retorno para Excel (sem mexer na FUP)", expanded=False):
    st.markdown(
        """
        Gera um **Excel novo** com relacionamento:

        - **Verde (pesquisa):** PO com Release + Número da linha  
        - **Amarelo (retorno):** Data da Promessa, Observações de Coleta, Número da NF  

        **Não altera** o `relatorio_fup.xlsm`. Serve para conferir o match e, se quiser,
        fazer *PROCV* / Power Query na planilha principal depois.
        """
    )
    if st.button("Gerar Excel de retorno", type="primary", key="btn_export_retorno_fup"):
        try:
            resultado = exportar_retorno_fup_excel(registros)
            caminho = Path(resultado["arquivo"])
            st.success(
                f"Arquivo gerado: **{caminho.name}** · "
                f"Match na FUP: **{resultado['encontrados']}** · "
                f"Sem match: **{resultado['nao_encontrados']}**"
            )
            st.download_button(
                label="⬇️ Baixar Excel de retorno",
                data=caminho.read_bytes(),
                file_name=caminho.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_retorno_fup",
            )
            st.caption(f"Também salvo na pasta do projeto: `{caminho.name}`")
        except Exception as exc:
            st.error(f"Erro ao gerar Excel: {exc}")

# ── Filtros ─────────────────────────────────────────────────────────────────
st.subheader("Filtros")

busca_geral = st.text_input(
    "🔎 Busca geral",
    placeholder="Procura em todos os campos...",
)

f1, f2, f3, f4 = st.columns(4)
with f1:
    filtro_nome = st.text_input("Nome")
with f2:
    filtro_email = st.text_input("Email")
with f3:
    filtro_po = st.text_input("PO com Release")
with f4:
    filtro_linha = st.text_input("Linha")

f5, f6, f7, f8 = st.columns(4)
with f5:
    filtro_nf = st.text_input("Número da NF")
with f6:
    usar_filtro_data = st.checkbox("Filtrar por Data da Promessa")
    filtro_data = None
    if usar_filtro_data:
        filtro_data = st.date_input("Data da Promessa", format="DD/MM/YYYY")
with f7:
    somente_incompletos = st.checkbox(
        "⚠️ Só incompletos",
        help="Exibe apenas registros sem NF ou sem observação de coleta.",
    )
    filtro_prazo = st.selectbox(
        "Prazo do form",
        ["Todos", "No prazo", "Atrasado", "Sem prazo", "Sem data"],
        help="Compara a data do envio com a data limite definida acima.",
    )
with f8:
    itens_pagina = st.selectbox("Registros por página", OPCOES_PAGINA, index=1)

registros_filtrados = aplicar_filtros(
    registros,
    filtro_nome,
    filtro_email,
    filtro_po,
    filtro_nf,
    filtro_linha,
    busca_geral,
    filtro_data,
    somente_incompletos,
    filtro_prazo,
)

total_paginas = max(1, (len(registros_filtrados) + itens_pagina - 1) // itens_pagina)

if "pagina_atual" not in st.session_state:
    st.session_state.pagina_atual = 1

st.session_state.pagina_atual = min(st.session_state.pagina_atual, total_paginas)

m1, m2 = st.columns(2)
m1.metric("Registros filtrados", len(registros_filtrados))
m2.metric("Página", f"{st.session_state.pagina_atual} de {total_paginas}")

# ── Paginação ───────────────────────────────────────────────────────────────
pg1, pg2, pg3 = st.columns([1, 2, 1])
with pg1:
    if st.button("← Anterior", disabled=st.session_state.pagina_atual <= 1):
        st.session_state.pagina_atual -= 1
        st.rerun()
with pg2:
    nova_pagina = st.number_input(
        "Ir para página",
        min_value=1,
        max_value=total_paginas,
        value=st.session_state.pagina_atual,
        step=1,
    )
    if nova_pagina != st.session_state.pagina_atual:
        st.session_state.pagina_atual = int(nova_pagina)
        st.rerun()
with pg3:
    if st.button("Próxima →", disabled=st.session_state.pagina_atual >= total_paginas):
        st.session_state.pagina_atual += 1
        st.rerun()

inicio = (st.session_state.pagina_atual - 1) * itens_pagina
fim = inicio + itens_pagina
registros_pagina = registros_filtrados[inicio:fim]

# ── Tabela ──────────────────────────────────────────────────────────────────
if registros_pagina:
    df = registros_para_dataframe(registros_pagina)
    st.dataframe(aplicar_destaque(df), use_container_width=True, hide_index=True)

    excel_bytes = gerar_excel(registros_filtrados)
    nome_arquivo = f"fornecedores_{agora.strftime('%Y%m%d_%H%M%S')}.xlsx"
    st.download_button(
        label="📥 Exportar Excel (todos filtrados)",
        data=excel_bytes,
        file_name=nome_arquivo,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
elif registros:
    st.info("Nenhum registro encontrado com os filtros aplicados.")
else:
    st.info("Nenhum registro encontrado. Envie respostas pelo formulário.")
