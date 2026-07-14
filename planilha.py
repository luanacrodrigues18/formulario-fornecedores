from __future__ import annotations

import json
import warnings
from datetime import date, datetime
from urllib.parse import quote
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


def _sem_avisos_openpyxl():
    return warnings.catch_warnings()


def _load_workbook(caminho: Path | str, **kwargs):
    with _sem_avisos_openpyxl():
        warnings.simplefilter("ignore", UserWarning)
        return load_workbook(caminho, **kwargs)

BASE_DIR = Path(__file__).resolve().parent
ARQUIVO_FUP = BASE_DIR / "relatorio_fup.xlsm"
ARQUIVO_CODIGOS = BASE_DIR / "fornecedores_codigos.json"
ARQUIVO_RESPOSTAS = BASE_DIR / "formulario_respostas.xlsx"

ABA_BASE = "Follow-up-Release"
ABA_FORM = "Form1"
ABA_FORM_ALT = "Formulario_Respostas"
LINHA_CABECALHO_BASE = 10
LINHA_INICIO_DADOS = 11
# Código do fornecedor vem de fornecedores_codigos.json (a planilha FUP não tem essa coluna).

COLUNAS_FORM = [
    "ID",
    "Hora de início",
    "Hora da conclusão",
    "Email",
    "Nome",
    "Número do PO com Release",
    "Data da Promessa",
    "Observações de Coleta",
    "Número da NF",
    "Informe o número da linha",
]

LARGURAS_COLUNAS = [8, 22, 22, 30, 25, 28, 18, 35, 18, 22]


def _normalizar_texto(valor: Any) -> str:
    if valor is None:
        return ""
    return str(valor).strip()


def _po_com_release(linha: tuple[Any, ...]) -> str:
    pedidos = _normalizar_texto(linha[30] if len(linha) > 30 else "")
    if pedidos and pedidos != "-":
        return pedidos

    acordo = _normalizar_texto(linha[0] if len(linha) > 0 else "")
    release = linha[1] if len(linha) > 1 else ""
    if acordo and release is not None and _normalizar_texto(release) != "":
        return f"{acordo}-{release}"
    return acordo


def _normalizar_codigo_fornecedor(valor: Any) -> str:
    texto = _normalizar_texto(valor)
    if not texto:
        return ""
    if texto.replace(".", "").isdigit():
        return texto.lstrip("0") or "0"
    return texto


def codigos_equivalentes(codigo: str) -> set[str]:
    base = _normalizar_codigo_fornecedor(codigo)
    bruto = _normalizar_texto(codigo)
    if not base and not bruto:
        return set()

    variantes = {base, bruto}
    if base.isdigit():
        variantes.add(base.zfill(10))
    return {v for v in variantes if v}


def mesmo_codigo_fornecedor(codigo_a: str, codigo_b: str) -> bool:
    return bool(codigos_equivalentes(codigo_a) & codigos_equivalentes(codigo_b))


def garantir_arquivo_codigos() -> None:
    if ARQUIVO_CODIGOS.is_file():
        return

    from database import (
        baixar_arquivo_storage,
        supabase_codigos_file,
        supabase_configurado,
        supabase_storage_bucket,
    )

    if not supabase_configurado():
        return

    try:
        baixar_arquivo_storage(supabase_codigos_file(), ARQUIVO_CODIGOS)
    except Exception as exc:
        mensagem = str(exc)
        if "11001" in mensagem or "getaddrinfo" in mensagem.lower():
            raise RuntimeError(
                "Não foi possível conectar ao Supabase para baixar o cadastro de códigos. "
                "Coloque fornecedores_codigos.json na pasta do projeto."
            ) from exc
        raise RuntimeError(
            f"Não foi possível baixar {supabase_codigos_file()} "
            f"do bucket {supabase_storage_bucket()}. "
            "Faça upload do arquivo no Supabase Storage."
        ) from exc

    if not ARQUIVO_CODIGOS.exists():
        raise RuntimeError(
            f"Não foi possível baixar {supabase_codigos_file()} "
            f"do bucket {supabase_storage_bucket()}."
        )
    carregar_mapa_codigos_manual.cache_clear()


@lru_cache(maxsize=1)
def carregar_mapa_codigos_manual() -> dict[str, str]:
    if not ARQUIVO_CODIGOS.is_file():
        return {}

    with ARQUIVO_CODIGOS.open(encoding="utf-8") as arquivo:
        dados = json.load(arquivo)

    mapa: dict[str, str] = {}
    if isinstance(dados, dict):
        for codigo, nome in dados.items():
            codigo_norm = _normalizar_codigo_fornecedor(str(codigo))
            nome_norm = _normalizar_texto(nome)
            if codigo_norm and nome_norm:
                mapa[codigo_norm] = nome_norm
    return mapa


def resolver_fornecedor_por_codigo(codigo: str) -> dict[str, str] | None:
    codigo = _normalizar_texto(codigo)
    if not codigo:
        return None

    for chave, nome in carregar_mapa_codigos_manual().items():
        if mesmo_codigo_fornecedor(chave, codigo):
            return {
                "codigo_fornecedor": _normalizar_codigo_fornecedor(chave),
                "fornecedor": nome,
            }

    for registro in carregar_base_fup():
        codigo_registro = registro.get("codigo_fornecedor", "")
        if codigo_registro and mesmo_codigo_fornecedor(codigo_registro, codigo):
            return {
                "codigo_fornecedor": _normalizar_codigo_fornecedor(codigo_registro),
                "fornecedor": registro["fornecedor"],
            }

    return None


def buscar_linhas_do_codigo(codigo: str) -> list[dict[str, Any]]:
    info = resolver_fornecedor_por_codigo(codigo)
    if not info:
        return []

    nome = info["fornecedor"]
    codigo_norm = info["codigo_fornecedor"]
    linhas = [r for r in carregar_base_fup() if r.get("fornecedor") == nome]
    com_codigo = [
        r
        for r in linhas
        if r.get("codigo_fornecedor")
        and mesmo_codigo_fornecedor(r["codigo_fornecedor"], codigo_norm)
    ]
    return com_codigo if com_codigo else linhas


def buscar_respostas_do_codigo(codigo: str, nome_fornecedor: str = "") -> list[dict[str, Any]]:
    from database import buscar_todos

    respostas: list[dict[str, Any]] = []
    for registro in buscar_todos():
        codigo_registro = registro.get("codigo_fornecedor", "")
        if codigo_registro and mesmo_codigo_fornecedor(codigo_registro, codigo):
            respostas.append(registro)
            continue
        if nome_fornecedor and _normalizar_texto(registro.get("nome", "")) == _normalizar_texto(
            nome_fornecedor
        ):
            respostas.append(registro)
    return respostas


def _linha_para_registro(indice: int, linha: tuple[Any, ...]) -> dict[str, Any]:
    return {
        "indice_base": indice,
        "acordo": _normalizar_texto(linha[0] if len(linha) > 0 else ""),
        "release": _normalizar_texto(linha[1] if len(linha) > 1 else ""),
        "numero_linha": _normalizar_texto(linha[6] if len(linha) > 6 else ""),
        "codigo_fornecedor": "",
        "fornecedor": _normalizar_texto(linha[11] if len(linha) > 11 else ""),
        "email_fornecedor": _normalizar_texto(linha[26] if len(linha) > 26 else ""),
        "numero_po_com_release": _po_com_release(linha),
        "data_promessa_base": linha[31] if len(linha) > 31 else None,
        "observacoes_base": _normalizar_texto(linha[32] if len(linha) > 32 else ""),
        "descricao_item": _normalizar_texto(linha[9] if len(linha) > 9 else ""),
    }


def _localizar_fup() -> Path | None:
    candidatos = [
        BASE_DIR / "relatorio_fup.xlsm",
        BASE_DIR / "Relatório - FUP.xlsm",
    ]
    for caminho in candidatos:
        if caminho.is_file():
            return caminho
    for caminho in BASE_DIR.glob("*.xlsm"):
        if "fup" in caminho.name.lower():
            return caminho
    return None


def garantir_arquivo_fup() -> None:
    global ARQUIVO_FUP

    local = _localizar_fup()
    if local:
        ARQUIVO_FUP = local
        return

    from database import baixar_fup_storage, supabase_configurado, supabase_fup_file, supabase_storage_bucket, supabase_url

    if not supabase_configurado():
        raise RuntimeError(
            "Planilha base não encontrada e Supabase não configurado. "
            "Coloque relatorio_fup.xlsm na pasta do projeto ou configure os Secrets."
        )

    try:
        baixar_fup_storage(ARQUIVO_FUP)
    except Exception as exc:
        mensagem = str(exc)
        if "11001" in mensagem or "getaddrinfo" in mensagem.lower():
            raise RuntimeError(
                "Não foi possível conectar ao Supabase (rede bloqueada ou URL incorreta). "
                f"URL configurada: {supabase_url()}. "
                "No PC, coloque relatorio_fup.xlsm na pasta do projeto."
            ) from exc
        raise

    if not ARQUIVO_FUP.exists():
        raise RuntimeError(
            f"Não foi possível baixar {supabase_fup_file()} do bucket {supabase_storage_bucket()}."
        )
    carregar_base_fup.cache_clear()
    carregar_mapa_codigos_manual.cache_clear()


@lru_cache(maxsize=1)
def carregar_base_fup() -> list[dict[str, Any]]:
    if not ARQUIVO_FUP.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {ARQUIVO_FUP}")

    with _sem_avisos_openpyxl():
        warnings.simplefilter("ignore", UserWarning)

        wb = _load_workbook(ARQUIVO_FUP, read_only=True, data_only=True)
        ws = wb[ABA_BASE]
        registros: list[dict[str, Any]] = []

        for indice, linha in enumerate(
            ws.iter_rows(min_row=LINHA_INICIO_DADOS, values_only=True),
            start=LINHA_INICIO_DADOS,
        ):
            if not linha or not any(linha):
                continue

            registro = _linha_para_registro(indice, linha)
            if registro["fornecedor"] and registro["numero_po_com_release"]:
                registros.append(registro)

        wb.close()

    return registros


def listar_fornecedores() -> list[str]:
    fornecedores = sorted({r["fornecedor"] for r in carregar_base_fup()})
    return fornecedores


def buscar_por_fornecedor(fornecedor: str) -> list[dict[str, Any]]:
    fornecedor = fornecedor.strip()
    return [r for r in carregar_base_fup() if r["fornecedor"] == fornecedor]


def buscar_por_po(numero_po: str) -> list[dict[str, Any]]:
    termo = numero_po.strip().lower()
    if not termo:
        return []

    return [
        r
        for r in carregar_base_fup()
        if termo in r["numero_po_com_release"].lower()
    ]


def buscar_por_po_e_linha(numero_po: str, numero_linha: str) -> list[dict[str, Any]]:
    po = numero_po.strip().lower()
    linha = numero_linha.strip()
    if not po or not linha:
        return []

    exatos = [
        r
        for r in carregar_base_fup()
        if r["numero_po_com_release"].lower() == po and r["numero_linha"] == linha
    ]
    if exatos:
        return exatos

    return [
        r
        for r in carregar_base_fup()
        if po in r["numero_po_com_release"].lower() and r["numero_linha"] == linha
    ]


def rotulo_linha(registro: dict[str, Any]) -> str:
    return (
        f"PO {registro['numero_po_com_release']} | "
        f"Linha {registro['numero_linha']} | "
        f"{registro['descricao_item'][:60]}"
    )


def _converter_data(valor: Any) -> date | None:
    if valor is None or _normalizar_texto(valor) in {"", "-"}:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    try:
        return datetime.fromisoformat(str(valor)).date()
    except ValueError:
        return None


def _linha_para_registro_resposta(row: tuple[Any, ...]) -> dict[str, Any]:
    if len(row) >= 10:
        numero_nf = row[8]
        numero_linha = row[9]
    else:
        numero_nf = ""
        numero_linha = row[8] if len(row) > 8 else ""

    return {
        "id": row[0],
        "hora_inicio": row[1],
        "hora_conclusao": row[2],
        "email": row[3],
        "nome": row[4],
        "numero_po_com_release": row[5],
        "data_promessa": row[6],
        "observacoes_coleta": row[7],
        "numero_nf": numero_nf,
        "numero_linha": numero_linha,
    }


def _aplicar_estilo_cabecalho(ws) -> None:
    fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")

    for col_idx, titulo in enumerate(COLUNAS_FORM, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, largura in enumerate(LARGURAS_COLUNAS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = largura

    ws.freeze_panes = "A2"


def _migrar_coluna_nf(ws) -> None:
    cabecalhos = [cell.value for cell in ws[1]]
    if not cabecalhos or "Número da NF" in cabecalhos:
        return

    ws.insert_cols(9)
    fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    cell = ws.cell(row=1, column=9, value="Número da NF")
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(9)].width = 18


def _nome_aba_respostas(sheetnames: list[str]) -> str | None:
    if ABA_FORM in sheetnames:
        return ABA_FORM
    if ABA_FORM_ALT in sheetnames:
        return ABA_FORM_ALT
    return None


def _garantir_arquivo_respostas() -> None:
    if ARQUIVO_RESPOSTAS.exists():
        wb = _load_workbook(ARQUIVO_RESPOSTAS)
        nome_aba = _nome_aba_respostas(wb.sheetnames)
        if nome_aba:
            ws = wb[nome_aba]
        else:
            ws = wb.active
            ws.title = ABA_FORM
        _migrar_coluna_nf(ws)
        wb.save(ARQUIVO_RESPOSTAS)
        wb.close()
        return

    wb = Workbook()
    ws = wb.active
    ws.title = ABA_FORM
    _aplicar_estilo_cabecalho(ws)
    wb.save(ARQUIVO_RESPOSTAS)
    wb.close()


def _proximo_id() -> int:
    ids: list[int] = []

    if ARQUIVO_RESPOSTAS.exists():
        wb = _load_workbook(ARQUIVO_RESPOSTAS, read_only=True, data_only=True)
        nome_aba = _nome_aba_respostas(wb.sheetnames)
        if nome_aba:
            ws = wb[nome_aba]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0] is not None:
                    try:
                        ids.append(int(row[0]))
                    except (TypeError, ValueError):
                        pass
        wb.close()

    if ARQUIVO_FUP.exists():
        wb = _load_workbook(ARQUIVO_FUP, read_only=True, data_only=True)
        if ABA_FORM in wb.sheetnames:
            ws = wb[ABA_FORM]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0] is not None:
                    try:
                        ids.append(int(row[0]))
                    except (TypeError, ValueError):
                        pass
        wb.close()

    return max(ids, default=0) + 1


def append_resposta_formulario(dados: dict[str, Any]) -> dict[str, Any]:
    _garantir_arquivo_respostas()

    registro_id = dados.get("id") or _proximo_id()
    hora_inicio = dados.get("hora_inicio")
    hora_conclusao = dados.get("hora_conclusao")
    data_promessa = dados.get("data_promessa")

    if isinstance(hora_inicio, str):
        hora_inicio = datetime.fromisoformat(hora_inicio)
    if isinstance(hora_conclusao, str):
        hora_conclusao = datetime.fromisoformat(hora_conclusao)
    if isinstance(data_promessa, str):
        data_promessa = datetime.fromisoformat(data_promessa).date()

    linha_excel = [
        registro_id,
        hora_inicio,
        hora_conclusao,
        dados.get("email", ""),
        dados.get("nome", ""),
        dados.get("numero_po_com_release", ""),
        data_promessa,
        dados.get("observacoes_coleta", ""),
        dados.get("numero_nf", ""),
        dados.get("numero_linha", ""),
    ]

    wb = _load_workbook(ARQUIVO_RESPOSTAS)
    nome_aba = _nome_aba_respostas(wb.sheetnames) or ABA_FORM
    ws = wb[nome_aba]
    ws.append(linha_excel)
    wb.save(ARQUIVO_RESPOSTAS)
    wb.close()

    return {
        "id": registro_id,
        "hora_inicio": hora_inicio,
        "hora_conclusao": hora_conclusao,
        "email": dados.get("email", ""),
        "nome": dados.get("nome", ""),
        "numero_po_com_release": dados.get("numero_po_com_release", ""),
        "data_promessa": data_promessa.isoformat() if data_promessa else "",
        "observacoes_coleta": dados.get("observacoes_coleta", ""),
        "numero_nf": dados.get("numero_nf", ""),
        "numero_linha": dados.get("numero_linha", ""),
    }


def _chave_po_linha(numero_po: str, numero_linha: str) -> tuple[str, str]:
    return str(numero_po or "").strip().lower(), str(numero_linha or "").strip().lower()


def resposta_existe(numero_po: str, numero_linha: str) -> dict[str, Any] | None:
    chave = _chave_po_linha(numero_po, numero_linha)
    for registro in carregar_respostas():
        if _chave_po_linha(
            registro.get("numero_po_com_release", ""),
            registro.get("numero_linha", ""),
        ) == chave:
            return registro
    return None


def montar_link_formulario(base_url: str, numero_po: str, numero_linha: str) -> str:
    base = base_url.rstrip("/")
    po = quote(str(numero_po).strip())
    linha = quote(str(numero_linha).strip())
    return f"{base}?po={po}&linha={linha}"


def carregar_respostas() -> list[dict[str, Any]]:
    if not ARQUIVO_RESPOSTAS.exists():
        return []

    wb = _load_workbook(ARQUIVO_RESPOSTAS, read_only=True, data_only=True)
    nome_aba = _nome_aba_respostas(wb.sheetnames)
    if not nome_aba:
        wb.close()
        return []

    ws = wb[nome_aba]
    registros: list[dict[str, Any]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        registros.append(_linha_para_registro_resposta(row))

    wb.close()
    return list(reversed(registros))


def data_promessa_inicial(registro: dict[str, Any]) -> date:
    data_base = _converter_data(registro.get("data_promessa_base"))
    return data_base or date.today()


def exportar_retorno_fup_excel(
    respostas: list[dict[str, Any]] | None = None,
    destino: Path | None = None,
) -> dict[str, Any]:
    """
    Gera um Excel NOVO (não altera relatorio_fup.xlsm) com:
      - Colunas verdes (pesquisa): PO com Release + Número da linha
      - Colunas amarelas (retorno): Data da Promessa, Observações, NF
      - Status do relacionamento com a aba Follow-up-Release
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    if respostas is None:
        from database import buscar_todos

        respostas = buscar_todos()

    garantir_arquivo_fup()
    carregar_base_fup.cache_clear()
    base_por_chave: dict[tuple[str, str], dict[str, Any]] = {}
    for registro_fup in carregar_base_fup():
        chave = _chave_po_linha(
            registro_fup.get("numero_po_com_release", ""),
            registro_fup.get("numero_linha", ""),
        )
        if chave[0] and chave[1] and chave not in base_por_chave:
            base_por_chave[chave] = registro_fup

    fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_amarelo = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_cab = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    font_cab = Font(bold=True, color="FFFFFF")

    cabecalhos = [
        ("Número do PO com Release", fill_verde),
        ("Informe o número da linha", fill_verde),
        ("Data da Promessa", fill_amarelo),
        ("Observações de Coleta", fill_amarelo),
        ("Número da NF", fill_amarelo),
        ("Match na FUP", fill_cab),
        ("Fornecedor (FUP)", fill_cab),
        ("Linha Excel FUP", fill_cab),
        ("ID resposta", fill_cab),
        ("Email", fill_cab),
        ("Nome", fill_cab),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Retorno_Formulario"

    for col_idx, (titulo, fill) in enumerate(cabecalhos, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.fill = fill if fill != fill_cab else fill_cab
        if col_idx <= 2:
            cell.fill = fill_verde
            cell.font = Font(bold=True, color="006100")
        elif col_idx <= 5:
            cell.fill = fill_amarelo
            cell.font = Font(bold=True, color="9C5700")
        else:
            cell.fill = fill_cab
            cell.font = font_cab
        cell.alignment = Alignment(horizontal="center", vertical="center")

    encontrados = 0
    nao_encontrados = 0

    for row_idx, reg in enumerate(respostas, start=2):
        chave = _chave_po_linha(
            reg.get("numero_po_com_release", ""),
            reg.get("numero_linha", ""),
        )
        fup = base_por_chave.get(chave)
        match = "Sim" if fup else "Não"
        if fup:
            encontrados += 1
        else:
            nao_encontrados += 1

        data_val = _converter_data(reg.get("data_promessa"))
        ws.cell(row=row_idx, column=1, value=_normalizar_texto(reg.get("numero_po_com_release", "")))
        ws.cell(row=row_idx, column=2, value=_normalizar_texto(reg.get("numero_linha", "")))
        ws.cell(row=row_idx, column=3, value=data_val or _normalizar_texto(reg.get("data_promessa", "")))
        ws.cell(row=row_idx, column=4, value=_normalizar_texto(reg.get("observacoes_coleta", "")))
        ws.cell(row=row_idx, column=5, value=_normalizar_texto(reg.get("numero_nf", "")))
        ws.cell(row=row_idx, column=6, value=match)
        ws.cell(row=row_idx, column=7, value=(fup or {}).get("fornecedor", ""))
        ws.cell(row=row_idx, column=8, value=(fup or {}).get("indice_base", ""))
        ws.cell(row=row_idx, column=9, value=reg.get("id", ""))
        ws.cell(row=row_idx, column=10, value=_normalizar_texto(reg.get("email", "")))
        ws.cell(row=row_idx, column=11, value=_normalizar_texto(reg.get("nome", "")))

    larguras = [28, 22, 18, 35, 18, 14, 40, 16, 12, 30, 25]
    for idx, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = largura
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cabecalhos))}{max(1, len(respostas) + 1)}"

    if destino is None:
        destino = BASE_DIR / f"fup_retorno_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(destino)

    return {
        "arquivo": str(destino),
        "total": len(respostas),
        "encontrados": encontrados,
        "nao_encontrados": nao_encontrados,
    }
